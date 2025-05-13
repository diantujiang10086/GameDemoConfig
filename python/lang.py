import openpyxl
import config
import os
import sys
class Lang:
    def __init__(self, index, title) -> None:
        self.index = index
        self.title = title
        self.list=[]
        pass

if __name__ == "__main__":
    if not os.path.exists(config.langXlsx):
        sys.exit()
    workbook = openpyxl.load_workbook(config.langXlsx, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    
    languages = []
    for i in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row = 1, column = i).value
        if value != None:
            languages.append(Lang(i, value))
    
    for row in range(2, worksheet.max_row + 1):
        key = worksheet.cell(row = row, column = 1).value
        if key == None:
            continue
        for language in languages:
            value = worksheet.cell(row = row, column = language.index).value
            language.list.append(key.lower() +'='+value)

    if not os.path.exists(config.langOutput):
        os.makedirs(config.langOutput)

    for language in languages:
        path = config.langOutput+"/"+language.title + ".txt"
        with open(path,"w+", encoding="utf-8") as f:
            f.write( '\n'.join(language.list))
            pass
        print('write: '+ path)