import os
import shutil
import openpyxl
from utls import getCellValue

cacheDir = 'cache'

def xlsxToCache(path):
    workbook = openpyxl.load_workbook(path)
    for name in workbook.sheetnames:
        if not name.isalpha():
            continue

        worksheet = workbook[name]
        tableLines = []
        for i in range(1, worksheet.max_row + 1):
            tableLine = []
            for j in range(1, worksheet.max_column + 1):
                cell = getCellValue(worksheet, i, j)
                tableLine.append(cell)
                
            tableLines.append('\t'.join(tableLine))

        if len(tableLines) > 0:
            savePath = os.path.join(cacheDir, name + ".csv")
            with open(savePath,"w", encoding="utf-8") as f:
                f.write('\n'.join(tableLines))
            print('generate cache: '+ savePath)


if __name__ == "__main__":

    shutil.rmtree(cacheDir)
    os.makedirs(cacheDir)

    for dir, _ ,files in os.walk("xlsxs"):
        for file in files:
            if file.endswith('xlsx'):
                path = os.path.join(dir, file)
                xlsxToCache(path)